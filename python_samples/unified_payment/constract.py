import openpyxl

wb = openpyxl.load_workbook('E:/TEST/订单明细_20210827_161051.xlsx')

sheet = wb.worksheets[0]

c = {}

for row in sheet.iter_rows():
    c[row[4].value] = row

sheet = wb.worksheets[1]
y = {}

for row in sheet.iter_rows():
    y[str(row[4].value)] = row


for e in y:
    if e in c.keys():
        target = y[e]
        source = c[e]

        target_pay_type = target[6].value
        source_pay_type = source[6].value

        if target_pay_type == '微信支付(统一支付)' and source_pay_type != '微信小程序支付':
            print('有星意订单号: ' + e + ', 支付方式: ' + target_pay_type + ', 支付时间: ' + str(target[31].value) + ', 统一支付订单号: ' + target[32].value + ', 统一支付方式: ' + source[6].value)
            
        
    else:
        print(e + ',' + y[e][6].value + ', 统一支付不存在该订单.')    

